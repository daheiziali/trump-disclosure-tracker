from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from analytics import asset_category, clean_asset_name, infer_ticker, is_public_investable_asset
from db import connect, init_db, record_health


PART6_LABEL = "Part 6: Other Assets and Income"
PART7_LABEL = "Part 7: Transactions"
HEADER_SCAN_ROWS = 30


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _cell(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def _find_latest_trump_annual(conn):
    return conn.execute(
        """
        SELECT f.id AS filing_id, f.source_document_id, f.person_name, f.filed_at
        FROM filings f
        JOIN source_documents sd ON sd.id = f.source_document_id
        WHERE f.filing_type = '278e'
          AND f.person_name = 'Trump, Donald J'
        ORDER BY COALESCE(f.filed_at, sd.published_at, sd.fetched_at) DESC, f.id DESC
        LIMIT 1
        """
    ).fetchone()


def _sheet_text(rows: list[tuple], limit: int = HEADER_SCAN_ROWS) -> str:
    return "\n".join(" ".join(_cell(c) for c in row if _cell(c)) for row in rows[:limit])


def _find_header(rows: list[tuple], required: set[str]) -> tuple[int, list[str]] | tuple[None, list[str]]:
    for idx, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        values = [_cell(c) for c in row]
        if required.issubset(set(values)):
            return idx, values
    return None, []


def _extract_page_number(rows: list[tuple]) -> int | None:
    for row in rows[:10]:
        text = " ".join(_cell(c) for c in row if _cell(c))
        match = re.search(r"Page\s+(\d+)\s+of\s+\d+", text, re.I)
        if match:
            return int(match.group(1))
    return None


def _is_account_row(values: list[str]) -> bool:
    return any(value.upper().startswith("INVESTMENT ACCOUNT #") for value in values)


def _account_name(values: list[str]) -> str | None:
    for value in values:
        if value.upper().startswith("INVESTMENT ACCOUNT #"):
            return value
    return None


def _is_number(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,6}", value))


def _format_date(value) -> str | None:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    text = _cell(value)
    if not text:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:\s+00:00:00)?", text):
        parsed = datetime.fromisoformat(text.split()[0])
        return f"{parsed.month}/{parsed.day}/{parsed.year}"
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
        month, day, year = text.split("/")
        if len(year) == 2:
            year = f"20{year}"
        return f"{int(month)}/{int(day)}/{year}"
    return text


def _normalize_tx_type(value: str) -> str | None:
    text = _cell(value).lower()
    if text == "purchase":
        return "Purchase"
    if text == "sale":
        return "Sale"
    if text == "exchange":
        return "Exchange"
    return None


def _raw(prefix: str, sheet: str, row_num: int, values: list[str]) -> str:
    return f"excel|{prefix}|{sheet}|row {row_num}|{' | '.join(values)}"


def parse_annual_excel(path: Path) -> dict[str, list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    holdings: list[dict] = []
    transactions: list[dict] = []
    sheet_summary: list[dict] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        text = _sheet_text(rows)
        page_num = _extract_page_number(rows)

        if PART6_LABEL in text:
            header_idx, headers = _find_header(rows, {"#", "Description", "Value", "Income Type", "Income Amount"})
            if header_idx is None:
                sheet_summary.append({"sheet": ws.title, "part": "Part 6", "rows": 0, "error": "missing_header"})
                continue
            count = 0
            account = None
            for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
                values = [_cell(c) for c in row]
                if not any(values):
                    continue
                if _is_account_row(values):
                    account = _account_name(values)
                    continue
                if len(values) < 6 or not _is_number(values[0]):
                    continue
                asset = clean_asset_name(values[1])
                if not asset:
                    continue
                ticker = infer_ticker(asset)
                holdings.append(
                    {
                        "section": "Part 6",
                        "account_name": account,
                        "asset_name": asset,
                        "ticker": ticker,
                        "value_range": values[3] or None,
                        "income_type": values[4] or None,
                        "income_range": values[5] or None,
                        "source_page": page_num,
                        "raw_text": _raw("Part 6", ws.title, row_idx, values[:6]),
                        "confidence": 0.99,
                        "asset_category": asset_category(asset, ticker),
                        "public_investable": is_public_investable_asset(asset, ticker),
                    }
                )
                count += 1
            sheet_summary.append({"sheet": ws.title, "part": "Part 6", "rows": count, "page": page_num})

        if PART7_LABEL in text:
            header_idx, headers = _find_header(rows, {"#", "Description", "Type", "Date", "Amount"})
            if header_idx is None:
                sheet_summary.append({"sheet": ws.title, "part": "Part 7", "rows": 0, "error": "missing_header"})
                continue
            count = 0
            account = None
            for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
                values = [_cell(c) for c in row]
                if not any(values):
                    continue
                if _is_account_row(values):
                    account = _account_name(values)
                    continue
                if len(row) < 5 or not _is_number(values[0]):
                    continue
                asset = clean_asset_name(values[1])
                tx_type = _normalize_tx_type(values[2])
                tx_date = _format_date(row[3])
                amount = values[4]
                if not asset or not tx_type or not tx_date or not amount:
                    continue
                ticker = infer_ticker(asset)
                transactions.append(
                    {
                        "section": "Part 7",
                        "account_name": account,
                        "asset_name": asset,
                        "ticker": ticker,
                        "transaction_type": tx_type,
                        "transaction_date": tx_date,
                        "amount_range": amount,
                        "source_page": page_num,
                        "raw_text": _raw("Part 7", ws.title, row_idx, [values[0], asset, tx_type, tx_date, amount]),
                        "confidence": 0.99,
                        "asset_category": asset_category(asset, ticker),
                        "public_investable": is_public_investable_asset(asset, ticker),
                    }
                )
                count += 1
            sheet_summary.append({"sheet": ws.title, "part": "Part 7", "rows": count, "page": page_num})

    return {"holdings": holdings, "transactions": transactions, "sheets": sheet_summary}


def _count_public(records: list[dict]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for record in records:
        category = record.get("asset_category") or "其他"
        counts[category] = counts.get(category, 0) + 1
    return counts


def ingest_annual_excel(path: Path, replace: bool = True) -> dict:
    init_db()
    parsed = parse_annual_excel(path)
    holdings = parsed["holdings"]
    transactions = parsed["transactions"]
    with connect() as conn:
        filing = _find_latest_trump_annual(conn)
        if not filing:
            raise RuntimeError("No Trump 278e filing found. Run OGE scan before Excel ingest.")
        if replace:
            conn.execute("DELETE FROM parsed_holdings WHERE source_document_id = ?", (filing["source_document_id"],))
            conn.execute("DELETE FROM parsed_transactions WHERE source_document_id = ?", (filing["source_document_id"],))

        inserted_holdings = 0
        inserted_transactions = 0
        for rec in holdings:
            cur = conn.execute(
                """
                INSERT OR IGNORE INTO parsed_holdings
                    (filing_id, source_document_id, person_name, section, account_name,
                     asset_name, ticker, value_range, income_type, income_range,
                     source_page, raw_text, confidence, review_state, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    filing["filing_id"],
                    filing["source_document_id"],
                    filing["person_name"],
                    rec["section"],
                    rec.get("account_name"),
                    rec["asset_name"],
                    rec.get("ticker"),
                    rec.get("value_range"),
                    rec.get("income_type"),
                    rec.get("income_range"),
                    rec.get("source_page"),
                    rec["raw_text"],
                    rec["confidence"],
                    "parsed",
                    _utc_now(),
                ),
            )
            inserted_holdings += cur.rowcount

        for rec in transactions:
            cur = conn.execute(
                """
                INSERT OR IGNORE INTO parsed_transactions
                    (filing_id, source_document_id, person_name, asset_name, ticker,
                     transaction_type, transaction_date, filed_date, amount_range,
                     source_page, raw_text, confidence, review_state, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    filing["filing_id"],
                    filing["source_document_id"],
                    filing["person_name"],
                    rec["asset_name"],
                    rec.get("ticker"),
                    rec["transaction_type"],
                    rec["transaction_date"],
                    filing["filed_at"],
                    rec["amount_range"],
                    rec.get("source_page"),
                    rec["raw_text"],
                    rec["confidence"],
                    "parsed",
                    _utc_now(),
                ),
            )
            inserted_transactions += cur.rowcount

        detail = (
            f"path={path}; holdings={len(holdings)}; transactions={len(transactions)}; "
            f"inserted_holdings={inserted_holdings}; inserted_transactions={inserted_transactions}"
        )
        record_health(conn, "annual_excel_parser", "ok", detail)

    return {
        "path": str(path),
        "replace": replace,
        "filing_id": filing["filing_id"],
        "source_document_id": filing["source_document_id"],
        "holdings": len(holdings),
        "transactions": len(transactions),
        "inserted_holdings": inserted_holdings,
        "inserted_transactions": inserted_transactions,
        "holding_categories": _count_public(holdings),
        "transaction_categories": _count_public(transactions),
        "public_holdings": sum(1 for row in holdings if row["public_investable"]),
        "public_transactions": sum(1 for row in transactions if row["public_investable"]),
        "sheets": parsed["sheets"],
    }


def main():
    parser = argparse.ArgumentParser(description="Parse Trump annual OGE 278e Excel tables.")
    parser.add_argument("path", type=Path, help="Path to Donald-J-Trump-2026-278ANNUAL.xlsx")
    parser.add_argument("--no-replace", action="store_true", help="Do not delete existing rows for the annual source document.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print counts without writing to SQLite.")
    args = parser.parse_args()

    if args.dry_run:
        parsed = parse_annual_excel(args.path)
        result = {
            "path": str(args.path),
            "holdings": len(parsed["holdings"]),
            "transactions": len(parsed["transactions"]),
            "public_holdings": sum(1 for row in parsed["holdings"] if row["public_investable"]),
            "public_transactions": sum(1 for row in parsed["transactions"] if row["public_investable"]),
            "holding_categories": _count_public(parsed["holdings"]),
            "transaction_categories": _count_public(parsed["transactions"]),
            "sheets": parsed["sheets"],
        }
    else:
        result = ingest_annual_excel(args.path, replace=not args.no_replace)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
