from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from analytics import asset_category, clean_asset_name, infer_ticker, is_public_investable_asset
from db import connect, init_db, record_health


DEFAULT_FILES = [
    "/Users/jiajingwen/Desktop/trump 披露/excel/0227DonaldJ.Trump2.26.2026278-T.pdf.xlsx",
    "/Users/jiajingwen/Desktop/trump 披露/excel/227DonaldJ.Trump2.26.2026278-T.pdf.xlsx",
    "/Users/jiajingwen/Desktop/trump 披露/excel/114Donald-J-Trump1.14.2026-278T.pdf.xlsx",
    "/Users/jiajingwen/Desktop/trump 披露/excel/06_29-Donald-J-Trump-06.25.2026-278T.pdf.xlsx",
    "/Users/jiajingwen/Desktop/trump 披露/excel/6_29-Donald-J-Trump-06.25.2026-278T.pdf.xlsx",
    "/Users/jiajingwen/Desktop/trump 披露/excel/5_12-TrumpDonaldJ.-05.08.2026-278T2.pdf.xlsx",
    "/Users/jiajingwen/Desktop/trump 披露/excel/5_12-TrumpDonaldJ.-05.08.2026-278T.pdf.xlsx",
    "/Users/jiajingwen/Desktop/trump 披露/excel/4_23Donald-J-Trump-4.20.2026-278T.pdf.xlsx",
]

FILE_TO_SOURCE_DOCUMENT_ID = {
    "06_29-Donald-J-Trump-06.25.2026-278T.pdf.xlsx": 4,
    "6_29-Donald-J-Trump-06.25.2026-278T.pdf.xlsx": 3,
    "5_12-TrumpDonaldJ.-05.08.2026-278T.pdf.xlsx": 8,
    "5_12-TrumpDonaldJ.-05.08.2026-278T2.pdf.xlsx": 9,
    "4_23Donald-J-Trump-4.20.2026-278T.pdf.xlsx": 11,
    "0227DonaldJ.Trump2.26.2026278-T.pdf.xlsx": 13,
    "227DonaldJ.Trump2.26.2026278-T.pdf.xlsx": 14,
    "114Donald-J-Trump1.14.2026-278T.pdf.xlsx": 16,
}

HEADER_ALIASES = {
    "asset": {"描述", "公司名称", "description"},
    "type": {"类型", "交易类型", "transaction_type"},
    "date": {"日期", "transaction_date"},
    "amount": {"金额", "金额范围", "amount_range"},
    "late": {"超过_30_天_通知", "超过30天通知", "notification_received_over_30_days_ago"},
}


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _cell(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def _format_date(value) -> str | None:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    text = _cell(value)
    if not text:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:\s+00:00:00)?", text):
        parsed = datetime.fromisoformat(text.split()[0])
        return f"{parsed.month}/{parsed.day}/{parsed.year}"
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
        month, day, year = text.split("/")
        if len(year) == 2:
            year = f"20{year}"
        return f"{int(month)}/{int(day)}/{year}"
    return text


def _normalize_type(value) -> str | None:
    text = _cell(value).lower()
    if "sale" in text or text == "卖出":
        return "Sale"
    if "purchase" in text or text == "买入":
        return "Purchase"
    if "exchange" in text or "换" in text:
        return "Exchange"
    return None


def _find_field(rows: list[tuple], names: set[str]) -> str | None:
    for row in rows[:12]:
        if len(row) < 2:
            continue
        key = _cell(row[0])
        if key in names:
            return _cell(row[1])
    return None


def _column_map(headers: list[str]) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(headers):
            if header in aliases:
                mapping[key] = idx
                break
    if {"asset", "type", "date", "amount"}.issubset(mapping):
        return mapping
    return None


def _raw(path: Path, sheet: str, row_num: int, values: list[str]) -> str:
    return f"excel278T|{path.name}|{sheet}|row {row_num}|{' | '.join(values)}"


def parse_transaction_excel(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    received_date = None
    records: list[dict] = []
    sheets: list[dict] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if ws.title.lower() == "fields":
            received_date = (
                _find_field(rows, {"收到_日期", "OGE收到日期", "OGE_接收日期", "oge_received_date"})
                or received_date
            )
            continue

        header_idx = None
        mapping = None
        for idx, row in enumerate(rows[:12]):
            headers = [_cell(c) for c in row]
            mapping = _column_map(headers)
            if mapping:
                header_idx = idx
                break
        if header_idx is None or mapping is None:
            continue

        count = 0
        for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            values = [_cell(c) for c in row]
            if not any(values):
                continue
            asset = clean_asset_name(values[mapping["asset"]])
            tx_type = _normalize_type(values[mapping["type"]])
            tx_date = _format_date(row[mapping["date"]])
            amount = values[mapping["amount"]]
            late_notice = values[mapping["late"]] if "late" in mapping and len(values) > mapping["late"] else None
            if not asset or not tx_type or not tx_date or not amount:
                continue
            ticker = infer_ticker(asset)
            records.append(
                {
                    "asset_name": asset,
                    "ticker": ticker,
                    "transaction_type": tx_type,
                    "transaction_date": tx_date,
                    "amount_range": amount,
                    "late_notice": late_notice,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                    "raw_text": _raw(path, ws.title, row_idx, [
                        asset,
                        tx_type,
                        tx_date,
                        late_notice or "",
                        amount,
                    ]),
                    "confidence": 0.99,
                    "asset_category": asset_category(asset, ticker),
                    "public_investable": is_public_investable_asset(asset, ticker),
                }
            )
            count += 1
        sheets.append({"sheet": ws.title, "rows": count})

    return {
        "path": str(path),
        "received_date": _format_date(received_date),
        "records": records,
        "sheets": sheets,
        "public_records": sum(1 for row in records if row["public_investable"]),
    }


def _filing_for_source_document(conn, source_document_id: int):
    return conn.execute(
        """
        SELECT f.id AS filing_id, f.person_name, f.filed_at, sd.id AS source_document_id
        FROM filings f
        JOIN source_documents sd ON sd.id = f.source_document_id
        WHERE sd.id = ? AND f.filing_type = '278-T'
        """,
        (source_document_id,),
    ).fetchone()


def ingest_transaction_excels(paths: list[Path], replace: bool = True) -> dict:
    init_db()
    results = []
    total_inserted = 0
    total_parsed = 0
    with connect() as conn:
        try:
            for path in paths:
                source_document_id = FILE_TO_SOURCE_DOCUMENT_ID.get(path.name)
                if not source_document_id:
                    raise RuntimeError(f"No source_document_id mapping for {path.name}")
                filing = _filing_for_source_document(conn, source_document_id)
                if not filing:
                    raise RuntimeError(f"No 278-T filing found for source_document_id={source_document_id}")
                parsed = parse_transaction_excel(path)
                records = parsed["records"]
                total_parsed += len(records)
                if replace:
                    conn.execute("DELETE FROM parsed_transactions WHERE source_document_id = ?", (source_document_id,))
                inserted = 0
                for rec in records:
                    cur = conn.execute(
                        """
                        INSERT OR IGNORE INTO parsed_transactions
                            (filing_id, source_document_id, person_name, asset_name, ticker,
                             transaction_type, transaction_date, filed_date, amount_range,
                             source_page, raw_text, confidence, review_state, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            filing["filing_id"],
                            source_document_id,
                            filing["person_name"],
                            rec["asset_name"],
                            rec.get("ticker"),
                            rec["transaction_type"],
                            rec["transaction_date"],
                            filing["filed_at"],
                            rec["amount_range"],
                            rec["source_row"],
                            rec["raw_text"],
                            rec["confidence"],
                            "parsed",
                            _utc_now(),
                        ),
                    )
                    inserted += cur.rowcount
                total_inserted += inserted
                results.append(
                    {
                        "path": str(path),
                        "source_document_id": source_document_id,
                        "filing_id": filing["filing_id"],
                        "received_date": parsed["received_date"],
                        "parsed": len(records),
                        "inserted": inserted,
                        "public_records": parsed["public_records"],
                        "sheets": parsed["sheets"],
                    }
                )
            record_health(
                conn,
                "transaction_excel_parser",
                "ok",
                f"files={len(paths)}; parsed={total_parsed}; inserted={total_inserted}",
            )
        except Exception as exc:
            record_health(conn, "transaction_excel_parser", "error", repr(exc))
            raise
    return {"files": len(paths), "parsed": total_parsed, "inserted": total_inserted, "results": results}


def main():
    parser = argparse.ArgumentParser(description="Parse converted 278-T Excel transaction files.")
    parser.add_argument("paths", nargs="*", type=Path, help="Excel files to parse. Defaults to current Trump 2026 files.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print counts without writing to SQLite.")
    parser.add_argument("--no-replace", action="store_true", help="Do not delete existing rows for mapped source documents.")
    args = parser.parse_args()
    paths = args.paths or [Path(p) for p in DEFAULT_FILES]
    if args.dry_run:
        result = {"files": len(paths), "results": []}
        result["parsed"] = 0
        result["public_records"] = 0
        for path in paths:
            parsed = parse_transaction_excel(path)
            result["parsed"] += len(parsed["records"])
            result["public_records"] += parsed["public_records"]
            result["results"].append({
                "path": str(path),
                "source_document_id": FILE_TO_SOURCE_DOCUMENT_ID.get(path.name),
                "received_date": parsed["received_date"],
                "parsed": len(parsed["records"]),
                "public_records": parsed["public_records"],
                "sheets": parsed["sheets"],
            })
    else:
        result = ingest_transaction_excels(paths, replace=not args.no_replace)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
